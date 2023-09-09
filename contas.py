from openpyxl import workbook
from datetime import datetime
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.worksheet.filters import (
    FilterColumn,
    CustomFilter,
    CustomFilters,
    DateGroupItem,
    Filters,
    )


planilha = load_workbook('contas.xlsx')
pagina= planilha['CONTAS A PAGAR']


pagina.cell(row=890, column=1, value='AGUA CLARA')

for rows in pagina.iter_rows(min_row=890, max_row= 897, max_col=11, values_only=True):
     print(rows)

